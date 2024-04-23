#!/usr/bin/python3
# -*- coding: utf-8 -*-
'''
@Author      :  ww1372247148@163.com
@AuthorDNS   :  wendirong.top
@CreateTime  :  2023-03-27 10:55:10
@FilePath    :  ExcelHandle.py
@FileVersion :  1.2
@LastEditTime:  2023-04-17
@FileDesc    :  用于读写Excel数据的Excel操作类
'''

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, GradientFill, PatternFill, Side, borders, colors, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import *


class ExcelFormatter:
    @staticmethod
    def get_KeyValMap_write(type_: int):
        '''
            获取字段名与列标的格式对照
            type_    : Enum(1: dingtalk_source)

        SUCCESS
        return dict         : 成功返回 dict对象

        ERROR
        return False        : 错误返回 False
        '''
        if type_ == 1:
            # dingtalk_source
            return {
                'id': (1, '唯一ID(id)'),
                'uid': (2, '序号(uid)'),
                'name': (3, '用户姓名(name)'),
                'ownGroup': (4, '所属组(ownGroup)'),
                'topGroup': (5, '一级部门(topGroup)'),
                'deptId': (6, '部门id(deptId)'),
                'department': (7, '部门(department)'),
                'email': (8, '邮箱(email)'),
                'position': (9, '职位(position)'),
                'jobNumber': (10, '工号(jobNumber)'),
                'entryTime': (11, '入职时间(entryTime)'),
                'nickName': (12, '昵称(nickName)'),
                'dingNum': (13, '钉钉号(dingNum)'),
                'remark': (14, '备注(remark)'),
            }

    @staticmethod
    def get_excelTitleMap(type_: int ):
        '''
            获取Excel输出数据的中英文对照的格式标题行
            type_    : Enum(1: dingtalk_source)

        SUCCESS
        return True         : 成功返回 True

        ERROR
        return False        : 错误返回 False
        '''
        if type_ == 1:
            # dingtalk_source
            return [
                '唯一ID(id)',
                '序号(uid)',
                '用户姓名(name)',
                '所属组(ownGroup)',
                '一级部门(topGroup)',
                '部门id(deptId)',
                '部门(department)',
                '邮箱(email)',
                '职位(position)',
                '工号(jobNumber)',
                '入职时间(entryTime)',
                '昵称(nickName)',
                '钉钉号(dingNum)',
                '备注(remark)',
            ]

    @staticmethod
    def write_merge_excelTitle(ws_: Worksheet, type_: int ):
        '''
            Excel写入标题行, 合并对应的单元格
            ws_      : Excel数据表
            type_    : Enum(1: dingtalk_source)

        SUCCESS
        return tuple        : 成功返回 tuple(coli, rowi)

        ERROR
        return False        : 错误返回 False
        '''
        titles = ExcelFormatter.get_excelTitleMap(type_)
        coli = 1
        rowi = 1
        for title_v in titles:
            if '#wrap' == title_v:
                coli = 1
                rowi += 1
                continue
            active_cell = get_column_letter(coli) + str(rowi)
            coli += 1
            if 'merge' in title_v:
                if 'start' in title_v:
                    ws_[active_cell] = title_v.split(':')[3]
                elif 'end' in title_v:
                    continue
                merge_cell_nw = title_v.split(':')[1]
                merge_cell_se = [v.split(':')[3] for v in titles if (f"{merge_cell_nw}:end") in v][0]
                ws_.merge_cells(f"{merge_cell_nw}:{merge_cell_se}")
            elif title_v != '':
                ws_[active_cell] = title_v
        return (coli, rowi)

    @staticmethod
    def format_excelTitle(ws_: Worksheet, format_cells: str, type_: int ):
        '''
            对Excel表格列头进行格式化
            ws_             : Excel数据表
            format_cells    : 需要进行格式化的表格列头
            type_           : Enum(1: dingtalk_source)

        SUCCESS
        return True         : 成功返回 True

        ERROR
        return False        : 错误返回 False
        '''
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        size12_font = Font(name='黑体', size=12, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color=colors.BLACK)
        fill_ffc000 = PatternFill(patternType='solid', fgColor='FFC000', bgColor='FFC000', fill_type=None, start_color=None, end_color=None)
        normal_side = Side(
            style=borders.BORDER_THIN,  # 边框样式，可选dashDot、dashDotDot、dashed、dotted、double、hair、medium、mediumDashDot、mediumDashDotDot、mediumDashed、slantDashDot、thick、thin
            color=colors.BLACK,  # 边框颜色，16进制rgb表示
        )
        for row in ws_[format_cells]:
            for cell in row:
                # 调整Excel格式布局: horizontal: center, vertical: center, wrap_text: True
                cell.alignment = center_alignment
                # 调整Excel格式字体: name: '黑体', size: 12, bold: False, italic: False, vertAlign: None, underline: 'none', strike: False, color: colors.BLACK
                cell.font = size12_font
                # 调整Excel格式填充: patternType: solid, fgColor: 'FFC000', bgColor: 'FFC000', fill_type: None, start_color: None, end_color: None
                cell.fill = fill_ffc000
                # 调整Excel格式边框: style: borders, BORDER_THIN, color: '000000'

                cell.border = Border(top=normal_side, bottom=normal_side, left=normal_side, right=normal_side)  # 上  # 下  # 左  # 右
                # 调整Excel单元格数据格式: number_format: FORMAT_TEXT
                cell.number_format = numbers.FORMAT_TEXT

        # 调整Excel格式: 行高列宽自定义
        if type_ == 1:
            # dingtalk_source
            for row_i in range(1, ws_.max_row + 1):
                ws_.row_dimensions[row_i].height = 36
            ws_.column_dimensions['A'].width = 22 + 0.63
            ws_.column_dimensions['B'].width = 6 + 0.63
            ws_.column_dimensions['C'].width = 28 + 0.63
            ws_.column_dimensions['D'].width = 50 + 0.63
            ws_.column_dimensions['E'].width = 22 + 0.63
            ws_.column_dimensions['F'].width = 36 + 0.63
            ws_.column_dimensions['G'].width = 50 + 0.63
            ws_.column_dimensions['H'].width = 22 + 0.63
            ws_.column_dimensions['I'].width = 22 + 0.63
            ws_.column_dimensions['J'].width = 22 + 0.63
            ws_.column_dimensions['K'].width = 10 + 0.63
            ws_.column_dimensions['L'].width = 10 + 0.63
            ws_.column_dimensions['M'].width = 10 + 0.63
            ws_.column_dimensions['N'].width = 10 + 0.63
            return True

    @staticmethod
    def format_excelCells(ws_: Worksheet, format_cells: str, type_: int ):
        '''
            对Excel表格数据进行格式化
            ws_             : Excel数据表
            format_cells    : 需要进行格式化的表格数据
            type_           : Enum(1: dingtalk_source)

        SUCCESS
        return True         : 成功返回 True

        ERROR
        return False        : 错误返回 False
        '''
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        fill_yellow = PatternFill(patternType='solid', fgColor='FFFF00', bgColor='FFFF00', fill_type=None, start_color=None, end_color=None)
        fill_00b0f0 = PatternFill(patternType='solid', fgColor='00B0F0', bgColor='00B0F0', fill_type=None, start_color=None, end_color=None)
        size11_font = Font(name='黑体', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color=colors.BLACK)
        normal_side = Side(
            style=borders.BORDER_THIN,  # 边框样式，可选dashDot、dashDotDot、dashed、dotted、double、hair、medium、mediumDashDot、mediumDashDotDot、mediumDashed、slantDashDot、thick、thin
            color=colors.BLACK,  # 边框颜色，16进制rgb表示
        )
        if type_ == 1:
            # dingtalk_source
            for row in ws_[format_cells]:
                # 调整Excel格式: 默认行高18磅, 多行则每行高15磅
                ws_.row_dimensions[row[0].row].height = 18 if len(row[3].value.split('\n')) == 0 else 15 * len(row[3].value.split('\n'))
                for cell in row:
                    # 调整Excel格式布局: horizontal: center, vertical: center, wrap_text: True
                    if cell.column_letter in ['B']:
                        cell.alignment = right_alignment
                    elif cell.column_letter in ['C']:
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = left_alignment
                    # 调整Excel格式字体: name: '黑体', size: 11, bold: False, italic: False, vertAlign: None, underline: 'none', strike: False, color: colors.BLACK
                    cell.font = size11_font
                    # 调整Excel格式边框: style: borders, BORDER_THIN, color: '000000'
                    cell.border = Border(top=normal_side, bottom=normal_side, left=normal_side, right=normal_side)  # 上  # 下  # 左  # 右
            return True


class ExcelStaticMethods:
    @staticmethod
    def getExcelLogger(prefix_: str = ''):
        logger_hander_ = CustomLogger(f"excel{prefix_}", start_in_log_=False)
        logger_ = logger_hander_.get_logs()
        return logger_, logger_hander_

    @staticmethod
    def writeExcel(write_json_: list, write_path_: str, wb_: Workbook, type_: int, *args, **kwargs):
        '''
            输出json数据到Excel文件
            write_json_ : 读取Json对象
            wb_         : 写入的Excel对象
            write_path_ : 写入的Excel对象完整路径
            type_       : Enum(1: dingtalk_source)
            *args       : 不定参数集. list列表允许输入多个参数
            **kwargs    : 不定参数集. dict集合允许输入多个键值对

        SUCCESS
        return True     : 成功返回 True

        ERROR
        return False    : 错误返回 False
        '''
        logger, logger_hander = ExcelStaticMethods.getExcelLogger()

        if len(wb_.sheetnames) == 1 and wb_.sheetnames[0] == 'Sheet':
            ws = wb_.active
        else:
            ws = wb_.create_sheet('NewSheet')

        type_list = [ 'dingtalk_source']
        title_list = ['钉钉通讯录']
        ws.title = title_list[type_ - 1]
        ws = ExcelStaticMethods.writeExcel_data(ws, write_json_, type_)

        logger.info('Success write {} ExcelFile. path: {}'.format(type_list[type_ - 1], write_path_))
        logger_hander.remove_logs()
        return True

    @staticmethod
    def writeExcel_data(write_ws_: Worksheet, write_json_: object, type_: int, *args, **kwargs):
        '''
            获取格式化后的Excel数据表
            Excel类型: type_
            write_ws_   : 需要写入的Worksheet对象
            write_json_ : 读取Json对象
            type_       : Enum(1: dingtalk_source)
            *args       : 不定参数集. list列表允许输入多个参数
            **kwargs    : 不定参数集. dict集合允许输入多个键值对

        SUCCESS
        return object   : 成功返回 Worksheet对象

        ERROR
        return None     : 错误返回 None
        '''
        coli, rowi = ExcelFormatter.write_merge_excelTitle(write_ws_, type_=type_)
        ExcelFormatter.format_excelTitle(write_ws_, 'A1:%s' % get_column_letter(coli - 1) + str(rowi), type_=type_)
        row_i = 1
        for row in write_json_ if type(write_json_) == list else write_json_.values():
            row['uid'] = row['uid'] if row.__contains__('uid') else row_i
            for col in range(1, write_ws_.max_column + 1):
                # 写入cell数据
                ret_col_key = [k for k, v in ExcelFormatter.get_KeyValMap_write(type_=type_).items() if v[0] == col][0]
                active_cell_value = row[ret_col_key] if ret_col_key in row.keys() else ''
                write_ws_[get_column_letter(col) + str(row_i + 1)] = '\n'.join([str(v) for v in active_cell_value]) if type(active_cell_value) == list else active_cell_value
            row_i += 1

        ExcelFormatter.format_excelCells(write_ws_, 'A2:%s' % get_column_letter(write_ws_.max_column) + str(write_ws_.max_row), type_=type_)
        return write_ws_
